# SPDX-FileCopyrightText: 2026 Mikołaj Kuranowski
# SPDX-License-Identifier: GPL-3.0-or-later

import re
from collections.abc import Iterable
from dataclasses import dataclass

from openpyxl.cell.cell import Cell as ExcelCell
from openpyxl.worksheet.worksheet import Worksheet as ExcelWorksheet

from . import extract
from .model import HeaderIndices, StationIndex


@dataclass
class Table:
    """Detected schedule table."""

    header: HeaderIndices
    stations: list[StationIndex]
    trains: range
    row_order: bool = False


def tables(ws: ExcelWorksheet) -> Iterable[Table]:
    """Detects and generates all schedule tables from the provided excel worksheet."""

    for anchor in anchors(ws):
        yield table(ws, anchor)


def table(ws: ExcelWorksheet, anchor: ExcelCell) -> Table:
    """Detects a table around the provided anchor (a cell with "열차번호")."""

    adjacent = extract.cell(ws, (anchor.row, anchor.column + 1))
    if not looks_like_train_number(adjacent):
        return row_table(ws, anchor)
    return column_table(ws, anchor)


def column_table(ws: ExcelWorksheet, anchor: ExcelCell) -> Table:
    """Detects a table around the provided anchor, where trips are oriented in columns."""

    # | 열차종별 | xxx   | xxx   | xxx   |
    # | 열차번호 | 100   | 102   | 103   |
    # |        | 00:00 | 00:00 | 00:00 |
    # | N      | 05:00 | 00:00 | 07:31 |
    # | N+1    | 05:05 | 06:15 | 00:00 | ⋯
    # | N+2    | 05:10 | 06:30 | 07:42 |
    # |        | 00:00 | 00:00 | 00:00 |
    # | 비고    |       |       |       |
    # | 종착역  | ...   | ...   | ...   |

    # Detect train rows
    start = anchor.column + 1
    end = ws.max_column + 1
    for col in range(start, end):
        value = extract.cell(ws, (anchor.row, col))
        if not looks_like_train_number(value):
            end = col
            break
    trains = range(start, end)

    # Detect the header
    train_number_row = anchor.row
    train_kind_row: None | int | str = None
    train_note_row: None | int | str = None
    station_rows = list[StationIndex]()

    for row in range(anchor.row - 1, ws.max_row + 1):
        coords = (row, anchor.column)
        value = extract.cell(ws, coords, extract.station_name)

        if "열차번호" in value:
            pass
        elif "종착역" in value and row > anchor.row:
            break
        elif "시발역" in value or "종착역" in value:
            pass
        elif "편성" in value or "열차종별" in value:
            train_kind_row = row
        elif re.search(r"비\s*고", value):
            train_note_row = row
        elif value:
            # Check if there might be a departure time row below
            next_row_value = extract.cell(ws, (row + 1, anchor.column))
            if not next_row_value and looks_like_departure_row(ws, row + 1, trains):
                is_first_station = len(station_rows) == 0
                station_rows.append(
                    StationIndex(
                        value,
                        row,
                        (1, 0),
                        allow_if_departure_only=is_first_station,
                    )
                )
            else:
                station_rows.append(StationIndex(value, row))

    # Ensure all columns were present, as a special case filling them for ITX-청춘
    is_itx_cheongchun = re.search(r"ITX-?청춘", ws.title) is not None
    if is_itx_cheongchun:
        train_kind_row = "ITX-청춘"

        if "평일" in ws.title:
            train_note_row = "평일"
        elif "휴일" in ws.title:
            train_note_row = "휴일"
        else:
            raise ValueError(f"unable to detect operating dates in {ws.title}")
    else:
        if train_kind_row is None:
            raise ValueError(f"no '열차종별' row in table {ws.title}.{anchor.coordinate}")
        if train_note_row is None:
            raise ValueError(f"no '비고' row in table {ws.title}.{anchor.coordinate}")

    # Return the detected table
    return Table(
        header=HeaderIndices(
            number=train_number_row,
            kind=train_kind_row,
            note=train_note_row,
        ),
        stations=station_rows,
        trains=trains,
        row_order=False,
    )


def row_table(ws: ExcelWorksheet, anchor: ExcelCell) -> Table:
    """Detects a table around the provided anchor, where trips are oriented in rows."""

    # | 열차번호 | N     | N+1   | N+2   |
    # | JUNK   | ...   |       |       | ⋯
    # | JUNK   | ...   |       |       |
    # | 1      | 05:00 | 00:00 | 05:15 |
    # With "편성"/"열차종별" and "비고" somewhere in the extra columns

    # Detect the header
    train_number_col = anchor.column
    train_kind_col: None | int = None
    train_note_col: None | int = None
    station_cols = list[StationIndex]()

    for col in range(anchor.column + 1, ws.max_column + 1):
        coords = (anchor.row, col)
        value = extract.cell(ws, coords, extract.station_name)
        if "열차번호" in value or not value:
            break  # extra "열차번호" in header would belong to a different table
        elif "편성" in value or "열차종별" in value:
            train_kind_col = col
        elif "비고" in value:
            train_note_col = col
        else:
            station_cols.append(StationIndex(value, col))

    # Ensure all columns were present
    if train_kind_col is None:
        raise ValueError(f"no '편성' column in table {ws.title}.{anchor.coordinate}")
    if train_note_col is None:
        raise ValueError(f"no '비고' column in table {ws.title}.{anchor.coordinate}")

    # Detect train rows
    start = -1
    end = ws.max_row + 1
    for row in range(anchor.row + 1, ws.max_row + 1):
        value = extract.cell(ws, (row, anchor.column))

        if looks_like_train_number(value):
            if start < 0:
                start = row
        else:
            if start > 0:
                end = row
                break

    # Return the detected table
    return Table(
        header=HeaderIndices(
            number=train_number_col,
            kind=train_kind_col,
            note=train_note_col,
        ),
        stations=station_cols,
        trains=range(start, end),
        row_order=True,
    )


def anchors(ws: ExcelWorksheet) -> Iterable[ExcelCell]:
    """Detects all anchors (cells with "열차번호") in the provided excel worksheet."""
    for row in ws:
        for cell in row:
            if "열차번호" in extract.string(cell.value):
                assert isinstance(cell, ExcelCell)
                yield cell


def looks_like_train_number(x: str) -> bool:
    """Returns True if the string is a plausible train number.

    >>> looks_like_train_number("42")
    True
    >>> looks_like_train_number("burger")
    False
    """
    return re.search(r"\d+", x) is not None


def looks_like_departure_row(ws: ExcelWorksheet, row: int, cols: range) -> bool:
    return any(extract.cell(ws, (row, col), extract.time) for col in cols)
