# SPDX-FileCopyrightText: 2026 Mikołaj Kuranowski
# SPDX-License-Identifier: GPL-3.0-or-later

import re
import sys
from abc import ABC, abstractmethod
from collections.abc import Generator, Iterable
from contextlib import closing, contextmanager
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from typing import IO, Any, NamedTuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell as ExcelCell
from openpyxl.workbook.workbook import Workbook as ExcelWorkbook
from openpyxl.worksheet.worksheet import Worksheet as ExcelWorksheet

KO_TRAIN_NUMBER = "열차번호"
KO_TRAIN_KIND = "열차종별"
KO_COMPOSITION = "편성"
KO_NOTE = "비고"
KO_FIRST_STATION = "시발역"
KO_LAST_STATION = "종착역"
KO_THROUGH = "경유"


@dataclass
class ScrapedStopTime:
    stop: str
    time: int

    def as_json(self) -> dict[str, Any]:
        return {"stop": self.stop, "time": self.time}


@dataclass
class ScrapedTrip:
    number: str
    kind: str
    note: str
    stops: list[ScrapedStopTime]

    def as_json(self) -> dict[str, Any]:
        return {
            "number": self.number,
            "kind": self.kind,
            "note": self.note,
            "stops": [i.as_json() for i in self.stops],
        }


def scrape_from_xlsx(path: str) -> Iterable[ScrapedTrip]:
    with closing(load_workbook(path)) as wb:
        yield from scrape_from_workbook(wb, path)


def scrape_from_workbook(wb: ExcelWorkbook, path: str = "") -> Iterable[ScrapedTrip]:
    for ws in wb.worksheets:
        if "보는방법" in ws.title:
            continue
        print(path, ws.title, file=sys.stderr)
        yield from scrape_from_worksheet(ws)


def scrape_from_worksheet(ws: ExcelWorksheet) -> Iterable[ScrapedTrip]:
    # Any single Worksheet can contain multiple (usually 2) tables,
    # either in the row-wise (KTX) or column-wise (standard) format.
    #
    # Both formats "anchor" around a cell with "열차번호" ("train number").
    #
    # The row-wise format looks like this:
    #
    # | 열차번호 | N     | N+1   | N+2   |
    # | JUNK   | ...   |       |       | ⋯
    # | JUNK   | ...   |       |       |
    # | 1      | 05:00 | 00:00 | 05:15 |
    #   ⋮
    #
    # With the following caveats:
    # 1. Extra columns "비고" (notice) and "편성" (train kind) are somewhere in N... columns,
    #    anything else being station names.
    # 2. Usually there are only exactly 2 "junk" rows below the train number,
    #    reserved for translated station names, the parser can handle anything between 0 and 5.
    # 3. Another table can be immediately to the right of the current one,
    #    and care must be taken not to mistakenly take another "열차번호" as a station name.
    #
    #
    # The column-wise format looks like this:
    #
    # | 열차종별 | xxx   | xxx   | xxx   |
    # | 열차번호 | 100   | 102   | 103   |
    # |        | 00:00 | 00:00 | 00:00 |
    # | N      | 05:00 | 00:00 | 07:31 |
    # | N+1    | 05:05 | 06:15 | 00:00 | ⋯
    # | N+2    | 05:10 | 06:30 | 07:42 |
    # |        | 00:00 | 00:00 | 00:00 |
    # | 비고    |       |       |       |
    # | 종착역  | ...   | ...   | ...   |
    #
    # With the following caveats
    # 1. Row-wise, the table starts at 열차번호, unless the row above is 열차종별.
    # 2. Row-wise, the table ends after seeing both 비고 and 종착역.
    # 3. Column-wise, the table starts at the column with 열차번호, and that column is treated
    #    as a "header" column.
    # 4. Column-wise, the table ends at the first column where 열차번호 doesn't contain any numbers.
    # 5. The header column can be in any order, with:
    #    - 열차번호 being the train number
    #    - 열차종별 being the train kind
    #    - 비고 being the note
    #    - 종착역 and empty cells being disregarded
    #    - anything else being treated as a station name, with parenthesis stripped if they enclose the name
    #
    # Other generic notes:
    # - The column-wise format can be identified by checking if the value to its right is a number.
    # - 00:00:00 represents "does not stop", which means there's no way to tell if a train stops perfectly at midnight.
    # - Times represent "clockface values", care must be taken when interpreting values past-midnight
    # - Some time cells contain "xx선경유", meaning "through to xxx line". Such values need to be ignored,
    #    but this also means that the same train can be generated multiple times with different stops,
    #    and such scraped trips must be recombined back into one train.

    for anchor in _find_table_anchors_in_worksheet(ws):
        if _is_column_wise_table(ws, anchor):
            yield from ColumnTableScraper(ws, anchor).extract_all()
        else:
            yield from RowTableScraper(ws, anchor).extract_all()


class _HeaderIndices(NamedTuple):
    number: int
    kind: int
    note: int


_StationIndices = list[tuple[int, str]]


class _TableScraper(ABC):
    ws: ExcelWorksheet
    header: _HeaderIndices
    stations: _StationIndices
    iter_range: range

    def __init__(self, ws: ExcelWorksheet, anchor: ExcelCell) -> None:
        assert KO_TRAIN_NUMBER in _stringify(anchor.value)
        self.ws = ws
        self.header, self.stations = self.extract_header(ws, anchor)
        self.iter_range = self.find_iteration_range(ws, anchor)

    @staticmethod
    @abstractmethod
    def extract_header(
        ws: ExcelWorksheet,
        anchor: ExcelCell,
    ) -> tuple[_HeaderIndices, _StationIndices]:
        raise NotImplementedError

    @staticmethod
    @abstractmethod
    def find_iteration_range(ws: ExcelWorksheet, anchor: ExcelCell) -> range:
        raise NotImplementedError

    def extract_all(self) -> Iterable[ScrapedTrip]:
        for idx in self.iter_range:
            if trip := self.extract(idx):
                yield trip

    @abstractmethod
    def extract(self, __index: int) -> ScrapedTrip | None:
        raise NotImplementedError


class RowTableScraper(_TableScraper):
    @staticmethod
    def extract_header(
        ws: ExcelWorksheet, anchor: ExcelCell
    ) -> tuple[_HeaderIndices, _StationIndices]:
        number = anchor.column  # The anchor always contains the train number
        kind = -1
        note = -1
        stations = _StationIndices()

        row = anchor.row
        for column in range(anchor.column + 1, ws.max_column + 1):
            # NOTE: we deliberately skip the anchor column, in order not to trip the
            #       `break` on adjacent-table detection (when a cell contains 열차번호).
            cell = ws.cell(row, column)
            value = _stringify(cell.value)

            if value == "" or KO_TRAIN_NUMBER in value:
                break
            elif KO_COMPOSITION in value or KO_TRAIN_KIND in value:
                kind = column
            elif KO_NOTE in value:
                note = column
            else:
                stations.append((column, _clean_station_name(value)))

        if kind < 0:
            raise ValueError(
                f"no '{KO_COMPOSITION}' column in table {ws.title}:{anchor.coordinate}"
            )
        if note < 0:
            raise ValueError(f"no '{KO_NOTE}' column in table {ws.title}:{anchor.coordinate}")

        return _HeaderIndices(number, kind, note), stations

    @staticmethod
    def find_iteration_range(ws: ExcelWorksheet, anchor: ExcelCell) -> range:
        start = -1

        column = anchor.column
        for row in range(anchor.row + 1, ws.max_row + 1):
            cell = ws.cell(row, column)
            value = _stringify(cell.value)

            if _looks_like_train_number(value):
                if start < 0:
                    start = row
            else:
                if start > 0:
                    return range(start, row)
        return range(start, ws.max_row + 1)

    def extract(self, row: int) -> ScrapedTrip:
        return ScrapedTrip(
            number=_stringify(self.ws.cell(row, self.header.number).value),
            kind=_stringify(self.ws.cell(row, self.header.kind).value),
            note=_extract_note(self.ws.cell(row, self.header.note).value),
            stops=[
                ScrapedStopTime(stop, time)
                for column, stop in self.stations
                # XXX: Deliberately ignore None and 0, as the latter is also used to mark "does not stop"
                if (time := _extract_time(self.ws.cell(row, column).value))
            ],
        )


class ColumnTableScraper(_TableScraper):
    @staticmethod
    def extract_header(
        ws: ExcelWorksheet, anchor: ExcelCell
    ) -> tuple[_HeaderIndices, list[tuple[int, str]]]:
        number = anchor.row  # The anchor always contains the train number
        kind = -1
        note = -1
        stations = _StationIndices()

        column = anchor.column
        for row in range(anchor.row - 1, ws.max_row + 1):
            cell = ws.cell(row, column)
            value = _stringify(cell.value)

            if KO_COMPOSITION in value or KO_TRAIN_KIND in value:
                kind = row
            elif KO_NOTE in value or re.fullmatch(r"비\s+고", value):
                note = row
            elif KO_LAST_STATION in value:
                break
            elif KO_TRAIN_NUMBER in value:
                pass
            elif value:
                stations.append((row, _clean_station_name(value)))
        else:
            raise ValueError(f"no '{KO_LAST_STATION}' row in table {ws.title}:{anchor.coordinate}")

        if kind < 0:
            raise ValueError(f"no '{KO_TRAIN_KIND}' row in table {ws.title}:{anchor.coordinate}")
        if note < 0:
            raise ValueError(f"no '{KO_NOTE}' row in table {ws.title}:{anchor.coordinate}")

        return _HeaderIndices(number, kind, note), stations

    @staticmethod
    def find_iteration_range(ws: ExcelWorksheet, anchor: ExcelCell) -> range:
        row = anchor.row
        for column in range(anchor.column + 1, ws.max_column + 1):
            cell = ws.cell(row, column)
            value = _stringify(cell.value)

            if not _looks_like_train_number(value):
                return range(anchor.column + 1, column)

        raise ValueError(f"unable to detect train columns in table {ws.title}:{anchor.coordinate}")

    def extract(self, column: int) -> ScrapedTrip | None:
        return ScrapedTrip(
            number=_stringify(self.ws.cell(self.header.number, column).value),
            kind=_stringify(self.ws.cell(self.header.kind, column).value),
            note=_extract_note(self.ws.cell(self.header.note, column).value),
            stops=[
                ScrapedStopTime(stop, time)
                for row, stop in self.stations
                # XXX: Deliberately ignore None and 0, as the latter is also used to mark "does not stop"
                if (time := _extract_time(self.ws.cell(row, column).value))
            ],
        )


def _find_table_anchors_in_worksheet(ws: ExcelWorksheet) -> Iterable[ExcelCell]:
    """Finds all cells which contain '열차번호' in the provided sheet."""
    for row in ws:
        for cell in row:
            if KO_TRAIN_NUMBER in _stringify(cell.value):
                assert isinstance(cell, ExcelCell), "MergedCell can't contain a string"
                yield cell


def _is_column_wise_table(ws: ExcelWorksheet, anchor: ExcelCell) -> bool:
    """Returns True if the table around the provided anchor looks like
    a column-wise schedule table. On False, the table looks like a row-wise table.
    """
    adjacent = ws.cell(anchor.row, anchor.column + 1)
    return _looks_like_train_number(_stringify(adjacent.value))


def _looks_like_train_number(x: str) -> bool:
    """Returns True if the string is a plausible train number.

    >>> looks_like_train_number("42")
    True
    >>> looks_like_train_number("burger")
    False
    """
    return re.search(r"\d+", x) is not None


def _clean_station_name(x: str) -> str:
    """Returns `x`, unless the value is enclosed in parentheses - in which case
    those are stripped.

    >>> clean_station_name("Foo")
    'Foo'
    >>> clean_station_name("(Bar)")
    'Bar'
    >>> clean_station_name("Spam (Eggs)")
    'Spam (Eggs)'
    """

    if x and x[0] == "(" and x[-1] == ")":
        return x[1:-1]
    return x


def _stringify(x: Any) -> str:
    """_stringify converts a cell value into a string. This is almost equivalent to `str(x)`,
    with the notable exception that `str(None)` given an empty string.

    >>> _stringify(42)
    '42'
    >>> _stringify("foo")
    'foo'
    >>> _stringify(None)
    ''
    >>> _stringify(datetime(2025, 12, 25, 14, 8))
    '2025-12-25 14:08:00'
    """

    match x:
        case None:
            return ""
        case str():
            return x.strip()
        case _:
            return str(x)


def _extract_time(x: Any) -> int | None:
    """Extracts a clockface time from a cell value, as seconds-since-midnight.

    - For datetime.time, it's simply `x.hour * 3600 + x.minute * 60 + x.second`.
    - For datetime.datetime, the behavior is the same as for datetime.time().
    - For strings, the first occurrence of `HH:MM` or `HH:MM:SS` is parsed, otherwise
        None is returned.
    - For None, None is returned.
    - For all other kinds, a TypeError is raised.
    """
    match x:
        case time() | datetime():
            return x.hour * 3600 + x.minute * 60 + x.second
        case timedelta():
            return round(x.total_seconds())
        case str():
            if m := re.search(r"([0-9]{1,2}):([0-9]{2})(?::([0-9]{2}))?", x):
                hour = int(m[1])
                minute = int(m[2])
                second = int(m[3]) if m[3] else 0
                return hour * 3600 + minute * 60 + second
            return None
        case None:
            return None
        case _:
            raise TypeError(
                f"don't know how to extract time from {x!r} (of type {type(x).__qualname__})"
            )


def _extract_note(x: Any) -> str:
    """Extracts a train note from a cell value.
    This is almost equivalent to `_stringify(x)`, except that values of temporal types
    cause the empty string to be returned.
    """
    match x:
        case time() | datetime() | timedelta():
            return ""
        case _:
            return _stringify(x)


@contextmanager
def _output_file(path: str | None) -> Generator[IO[str], None, None]:
    """Context-manager wrapper that opens a file at the provided path
    for writing, unless that path is None or exactly `-`, in which case sys.stdout is returned.
    """
    match path:
        case "-" | None:
            yield sys.stdout
        case _:
            with open(path, "w", encoding="utf-8") as f:
                yield f


if __name__ == "__main__":
    import argparse
    import json

    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument(
        "-o",
        "--output",
        metavar="JSONL_PATH",
        help="path to output .jsonl file",
    )
    arg_parser.add_argument(
        "input",
        nargs="+",
        metavar="XLSX_PATH",
        help="path to input .xlsx files",
    )
    args = arg_parser.parse_args()

    with _output_file(args.output) as f:
        for input_filename in args.input:
            for train in scrape_from_xlsx(input_filename):
                json.dump(train.as_json(), f, ensure_ascii=False)
                f.write("\n")
